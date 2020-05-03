import openpyxl

workbook = openpyxl.load_workbook('../../../data/test1.xlsx')
sheet = workbook["Sheet1"]

for rows in sheet.iter_rows(min_row=sheet.min_row, min_col=sheet.min_column, max_row=sheet.max_row, max_col=sheet.max_column):
    for cell in rows:
        print(cell.value)